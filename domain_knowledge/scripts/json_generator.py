#!/usr/bin/env python3
"""
GPP Domain Knowledge JSON Generator

This script converts GPP domain knowledge from Excel format to JSON files
that can be consumed by the Java GPP library.

Input:  Excel file with GPP criteria, documents, and patch data
Output: JSON files in the library's resources directory
"""

import openpyxl
import json
import os


# Configuration
INPUT_FILE_PATH_DOMAIN_KNOWLEDGE = '../sources/domain_knowledge.xlsx'
OUTPUT_DIR = '../../src/main/resources/domain_knowledge'
OUTPUT_FILE_NAME_GPP_CRITERIA = 'gpp_criteria.json'
OUTPUT_FILE_NAME_GPP_CRITERIA_DOCS = 'gpp_criteria_docs.json'
OUTPUT_FILE_NAME_GPP_PATCHES_DATA = 'gpp_patches_data.json'


def load_workbook(filepath):
    """Load Excel workbook with error handling."""
    try:
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        return workbook
    except FileNotFoundError:
        print(f"Error: The file '{filepath}' was not found.")
        return None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None


class GppCriteriaDoc:
    """Represents a GPP criteria document with metadata."""

    def __init__(self, row):
        self.name = row[0].value
        self.source = row[1].value
        self.document_reference = row[2].value
        self.publication_date = row[3].value
        self.relevant_cpv_codes = []

        # Parse CPV codes from various formats
        if row[4].value is not None:
            if isinstance(row[4].value, str):
                if ',' in row[4].value:
                    self.relevant_cpv_codes = [code.strip() for code in row[4].value.split(',')]
                else:
                    self.relevant_cpv_codes = [row[4].value.strip()]
            elif isinstance(row[4].value, (int, float)):
                self.relevant_cpv_codes = [str(int(row[4].value))]

        self.summary = row[5].value

    def to_dict(self):
        """Convert to dictionary for JSON serialization."""
        return {
            'name': self.name,
            'source': self.source,
            'documentReference': self.document_reference,
            'publicationDate': self.publication_date.isoformat() if self.publication_date else None,
            'relevantCpvCodes': self.relevant_cpv_codes,
            'summary': self.summary
        }


def load_gpp_criteria_docs(sheet):
    """Load GPP criteria documents from Excel sheet."""
    gpp_criteria_docs = []
    first_row = True

    for row in sheet.iter_rows():
        if first_row:
            first_row = False
            continue
        if row[0].value is None:
            break

        gpp_criteria_doc = GppCriteriaDoc(row)
        gpp_criteria_docs.append(gpp_criteria_doc)

    return gpp_criteria_docs


class GppCriterion:
    """Represents an individual GPP criterion."""

    def __init__(self, row):
        self.gpp_document = row[0].value.strip() if row[0].value else None
        self.gpp_source = row[1].value.strip() if row[1].value else None
        self.category = row[2].value.strip() if row[2].value else None
        self.criterion_type = row[3].value.strip() if row[3].value else None
        self.ambition_level = row[4].value.strip() if row[4].value else None
        self.id = row[5].value.strip() if row[5].value else None
        self.name = row[6].value.strip() if row[6].value else None

        # Parse CPV codes
        self.relevant_cpv_codes = []
        if row[7].value is not None:
            if isinstance(row[7].value, str):
                if ',' in row[7].value:
                    self.relevant_cpv_codes = [code.strip() for code in row[7].value.split(',')]
                else:
                    self.relevant_cpv_codes = [row[7].value.strip()]
            elif isinstance(row[7].value, (int, float)):
                self.relevant_cpv_codes = [str(int(row[7].value))]

        self.environmental_impact_type = row[8].value.strip() if row[8].value else None
        self.description = row[9].value.strip() if row[9].value else None
        self.selection_criterion_type = row[10].value.strip() if row[10].value else None

    def to_dict(self):
        """Convert to dictionary for JSON serialization."""
        return {
            'gppDocument': self.gpp_document,
            'gppSource': self.gpp_source,
            'category': self.category,
            'criterionType': self.criterion_type,
            'ambitionLevel': self.ambition_level,
            'id': self.id,
            'name': self.name,
            'relevantCpvCodes': self.relevant_cpv_codes,
            'environmentalImpactType': self.environmental_impact_type,
            'description': self.description,
            'selectionCriterionType': self.selection_criterion_type,
        }


def load_gpp_criteria(sheet):
    """Load GPP criteria from Excel sheet."""
    gpp_criteria = []
    first_row = True

    for row in sheet.iter_rows():
        if first_row:
            first_row = False
            continue
        if row[0].value is None:
            break

        gpp_criterion = GppCriterion(row)
        gpp_criteria.append(gpp_criterion)

    return gpp_criteria


class GppPatchData:
    """Represents eForm field mapping data for GPP integration."""

    def __init__(self, row):
        self.name = row[0].value.strip() if row[0].value else None

        # Parse BT IDs
        self.bt_ids = []
        if row[1].value is not None and row[1].value.strip() != '-':
            if ',' in row[1].value:
                self.bt_ids = [id.strip() for id in row[1].value.split(',')]
            else:
                self.bt_ids = [row[1].value.strip()]

        self.depends_on = None
        if row[2].value is not None and row[2].value.strip() != '-':
            self.depends_on = row[2].value.strip()

        self.path_in_lot = ''
        if row[3].value is not None and row[3].value.strip() != '-':
            self.path_in_lot = row[3].value.strip()

        self.value = None
        if row[4].value is not None and row[4].value.strip() != '-':
            self.value = row[4].value.strip()

    def to_dict(self):
        """Convert to dictionary for JSON serialization."""
        return {
            'name': self.name,
            'btIds': self.bt_ids,
            'dependsOn': self.depends_on,
            'pathInLot': self.path_in_lot,
            'value': self.value
        }


def load_gpp_patches_data(sheet):
    """Load GPP patch data from Excel sheet."""
    gpp_patches_data = []
    first_row = True

    for row in sheet.iter_rows():
        if first_row:
            first_row = False
            continue
        if row[0].value is None:
            break

        gpp_patch_data = GppPatchData(row)
        gpp_patches_data.append(gpp_patch_data)

    return gpp_patches_data


def main():
    """Main function to process Excel data and generate JSON files."""
    print("Loading domain knowledge from Excel...")

    workbook = load_workbook(INPUT_FILE_PATH_DOMAIN_KNOWLEDGE)
    if workbook is None:
        return

    print(f"Available sheets: {workbook.sheetnames}")

    # Load data from sheets
    print("Loading GPP criteria documents...")
    gpp_criteria_docs = load_gpp_criteria_docs(workbook['GPP Criteria Docs'])

    print("Loading GPP criteria...")
    gpp_criteria = load_gpp_criteria(workbook['GPP Criteria ENG'])

    print("Loading GPP patch data...")
    gpp_patches_data = load_gpp_patches_data(workbook['Patches'])

    # Ensure output directory exists
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Generate JSON files
    print("Generating JSON files...")

    # GPP Criteria Documents
    gpp_criteria_docs_json = [doc.to_dict() for doc in gpp_criteria_docs]
    with open(os.path.join(OUTPUT_DIR, OUTPUT_FILE_NAME_GPP_CRITERIA_DOCS), 'w', encoding='utf-8') as f:
        json.dump(gpp_criteria_docs_json, f, indent=2, ensure_ascii=False)
    print(f"✓ Created {OUTPUT_FILE_NAME_GPP_CRITERIA_DOCS}")

    # GPP Criteria
    gpp_criteria_json = [criterion.to_dict() for criterion in gpp_criteria]
    with open(os.path.join(OUTPUT_DIR, OUTPUT_FILE_NAME_GPP_CRITERIA), 'w', encoding='utf-8') as f:
        json.dump(gpp_criteria_json, f, indent=2, ensure_ascii=False)
    print(f"✓ Created {OUTPUT_FILE_NAME_GPP_CRITERIA}")

    # GPP Patches Data
    gpp_patches_data_json = [patch.to_dict() for patch in gpp_patches_data]
    with open(os.path.join(OUTPUT_DIR, OUTPUT_FILE_NAME_GPP_PATCHES_DATA), 'w', encoding='utf-8') as f:
        json.dump(gpp_patches_data_json, f, indent=2, ensure_ascii=False)
    print(f"✓ Created {OUTPUT_FILE_NAME_GPP_PATCHES_DATA}")

    print(f"\nSuccessfully generated {len(gpp_criteria_docs)} GPP documents, "
          f"{len(gpp_criteria)} GPP criteria, and {len(gpp_patches_data)} patch mappings.")


if __name__ == "__main__":
    main()
